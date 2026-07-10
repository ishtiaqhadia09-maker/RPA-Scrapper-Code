"""Post-process IQVIA Excel exports — add formula worksheets beside C/O/M."""

from __future__ import annotations

import csv
import logging
import re
import shutil
import tempfile
import zipfile
from collections.abc import Iterator
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

MTH_HEADER = re.compile(r"^MTH\d+$", re.I)
MTH_DATE_HEADER = re.compile(r"^MTH\s+\d{4}/\d{1,2}$", re.I)
MAT_DATE_HEADER = re.compile(r"^MAT\s+\d{4}/\d{1,2}$", re.I)
PERIOD_DATE_HEADER = re.compile(r"^(?:MTH|MAT)\s+\d{4}/\d{1,2}$", re.I)
EXCEL_SHEET_NAME_MAX = 31

RAW_FOR_NEXT_STAGE_HEADERS = [
    "MEASURES",
    "MAT1",
    "MAT13",
    "YTD1",
    "YTD13",
    "MTH1",
    "MTH13",
    "MTH14",
    "MTH2",
    "MTH3",
    "MTH4",
    "MTH5",
    "MTH6",
    "MTH7",
    "MTH8",
    "MTH9",
    "MTH10",
    "MTH11",
    "MTH12",
]

ATTRIBUTE_VALUE_SHEET_TITLE = "TABLE 1"
ATTRIBUTE_VALUE_MTH_HEADERS = [
    "MTH14",
    "MTH2",
    "MTH3",
    "MTH4",
    "MTH5",
    "MTH6",
    "MTH7",
    "MTH8",
    "MTH9",
    "MTH10",
    "MTH11",
    "MTH12",
]
ATTRIBUTE_VALUE_ATTRIBUTE_LABELS = ["MAT1", "MAT13", "YTD1", "YTD13", "MTH1", "MTH13"]
ATTRIBUTE_VALUE_HEADERS = (
    ["Pack1", "Pack2", "Brick", "Type", "MEASURES"]
    + ATTRIBUTE_VALUE_MTH_HEADERS
    + ["Attribute", "Value"]
)
# VTSTACK TYPE from source pivot sheet.
VTSTACK_TYPE_OWN = "OWN"
VTSTACK_TYPE_COMPETITOR = "COMPETITOR"
VTSTACK_TYPE_MARKET = "Market"
TABLE1_CSV_SUFFIX = "_TABLE1.csv"
VTSTACK_CSV_SUFFIX = "_VTSTACK.csv"
RAW_CSV_SUFFIX = "_RAW.csv"
CSV_BATCH_SIZE = 50_000
MAX_WORKBOOK_MB = 100


def combined_csv_path_for(xlsx_path: Path) -> Path:
    """Single CSV deliverable — all workbook sheets, one file."""
    return Path(xlsx_path).with_suffix(".csv")


def table1_csv_path_for(xlsx_path: Path) -> Path:
    """Sidecar CSV path for TABLE 1, e.g. CRESCOR_2026-07-07_TABLE1.csv."""
    path = Path(xlsx_path)
    return path.with_name(f"{path.stem}{TABLE1_CSV_SUFFIX}")


def vtstack_csv_path_for(xlsx_path: Path) -> Path:
    path = Path(xlsx_path)
    return path.with_name(f"{path.stem}{VTSTACK_CSV_SUFFIX}")


def raw_csv_path_for(xlsx_path: Path) -> Path:
    path = Path(xlsx_path)
    return path.with_name(f"{path.stem}{RAW_CSV_SUFFIX}")


def _num(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def _cell_at_row(row: tuple, col_1based: int):
    idx = col_1based - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _brick_code_only(value: object) -> str | None:
    text = _cell_text(value)
    if not text:
        return None
    # Expect formats like "1002 A.H.ISPHANI ROAD" or "1002  A.H..." → keep leading code.
    match = re.match(r"^\s*(\d+)", text)
    return match.group(1) if match else text.split()[0]


def _has_mkt_suffix(value: object) -> bool:
    """True when a pack/market label ends with MKT/MK/Market (e.g. CRESCOR DRV MKT)."""
    text = _cell_text(value).strip().upper()
    if not text:
        return False
    return (
        text.endswith("MKT")
        or text.endswith(" MKT")
        or text.endswith("MK")
        or text.endswith(" MK")
        or text.endswith("MARKET")
        or text.endswith(" MARKET")
    )


def _table1_type_label(
    *,
    pack1: str | None,
    pack2: str | None,
) -> str:
    """Type: Market (MKT/MK suffix), OWN (Pack1==Pack2), else COMPETITOR."""
    for candidate in (pack1, pack2):
        if candidate and _has_mkt_suffix(candidate):
            return VTSTACK_TYPE_MARKET
    if pack1 and pack2 and pack1 == pack2:
        return VTSTACK_TYPE_OWN
    return VTSTACK_TYPE_COMPETITOR


def _vtstack_row_source_type(row_tuple: tuple, type_col: int | None) -> str:
    if not type_col:
        return ""
    return _cell_text(_cell_at_row(row_tuple, type_col)).strip()


def _table1_pack2(
    *,
    row_tuple: tuple,
    key: tuple[str, str],
    type_col: int | None,
    pack_col: int | None,
    market_col: int | None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
) -> str | None:
    """TABLE 1 Pack2: C-pack, then O-pack, then M-market — not the shared market label on C/O."""
    source_type = _vtstack_row_source_type(row_tuple, type_col)
    normalized = source_type.upper()

    if normalized == VTSTACK_TYPE_COMPETITOR:
        return pack2_map.get(key)
    if normalized == VTSTACK_TYPE_OWN:
        pack = pack1_map.get(key)
        if pack:
            return pack
        if pack_col:
            return _cell_text(_cell_at_row(row_tuple, pack_col)).strip() or None
        return None
    if normalized in {VTSTACK_TYPE_MARKET.upper(), "MARKET"} or source_type == VTSTACK_TYPE_MARKET:
        if market_col:
            return _cell_text(_cell_at_row(row_tuple, market_col)).strip() or None
        return None

    if pack_col:
        pack = _cell_text(_cell_at_row(row_tuple, pack_col)).strip() or None
        if pack:
            return pack
    return pack2_map.get(key) or pack1_map.get(key)


def _vtstack_type_for_sheet(
    source_sheet: str | None,
    *,
    o_sheet: str | None,
    c_sheet: str | None,
    m_sheet: str | None,
) -> str:
    if source_sheet and o_sheet and source_sheet == o_sheet:
        return VTSTACK_TYPE_OWN
    if source_sheet and c_sheet and source_sheet == c_sheet:
        return VTSTACK_TYPE_COMPETITOR
    if source_sheet and m_sheet and source_sheet == m_sheet:
        return VTSTACK_TYPE_MARKET
    return VTSTACK_TYPE_MARKET


def _add_pack_map_entry(
    pack_map: dict[tuple[str, str], str],
    row_tuple: tuple,
    *,
    pack_col: int | None,
    brick_col: int | None,
    measures_col: int | None,
) -> None:
    if not pack_col or not brick_col or not measures_col:
        return
    brick_code = _brick_code_only(_cell_at_row(row_tuple, brick_col))
    if not brick_code:
        return
    measures = _cell_text(_cell_at_row(row_tuple, measures_col)).strip()
    pack = _cell_text(_cell_at_row(row_tuple, pack_col)).strip()
    if not measures or not pack:
        return
    pack_map[(brick_code, measures.lower())] = pack


def _build_pack_map_for_sheet(
    ws,
    *,
    sheet_name: str,
) -> dict[tuple[str, str], str]:
    """
    Return mapping (brick_code, measures) -> pack for one pivot sheet.

    - Uses detected header row (Product/Pack/Brick/Measures/MTH..)
    - Only keeps Units/Values rows
    """
    header_row, headers, header_index, measures_col = _scan_sheet_layout(ws)
    if header_row is None or not headers:
        logger.warning("Pack map — skipping %r (no header row)", sheet_name)
        return {}

    pack_col = header_index.get("pack")
    brick_col = header_index.get("brick")
    if not pack_col or not brick_col:
        logger.warning(
            "Pack map — skipping %r (missing pack/brick columns)", sheet_name
        )
        return {}

    header_keys = {_normalize_header_key(h) for h in headers if h}
    out: dict[tuple[str, str], str] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_tuple = tuple(row)
        if _is_repeated_header_row_tuple(
            row_tuple,
            measures_col=measures_col,
            header_keys=header_keys,
        ):
            continue
        if not _is_measure_data_row_tuple(row_tuple, measures_col):
            continue

        brick_raw = _cell_at_row(row_tuple, brick_col)
        brick_code = _brick_code_only(brick_raw)
        if not brick_code:
            continue

        measures = _cell_text(_cell_at_row(row_tuple, measures_col)).strip()
        if not measures:
            continue

        pack = _cell_text(_cell_at_row(row_tuple, pack_col)).strip()
        if not pack:
            continue

        out[(brick_code, measures.lower())] = pack

    logger.info("Pack map — %r: %d brick+measure rows", sheet_name, len(out))
    return out


def _mth_cols_from_numbered_headers(headers: list[str]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for col_idx, header in enumerate(headers, start=1):
        text = _cell_text(header).upper()
        if MTH_HEADER.match(text):
            cols[text] = col_idx
    return cols


def _calendar_from_header_cells(headers: list) -> dict[str, tuple[int, int]]:
    date_cols = [
        (col_idx, _cell_text(header))
        for col_idx, header in enumerate(headers, start=1)
        if _cell_text(header) and MTH_DATE_HEADER.match(_cell_text(header))
    ]
    return _build_mth_calendar_from_dates(date_cols) if date_cols else {}


def _column_layout_from_vtstack_headers(
    headers: list,
) -> tuple[int | None, dict[str, int], dict[str, tuple[int, int]]]:
    """
    Column indices for MEASURES / MTHn on a VTSTACK header row.

    VTSTACK inserts TYPE after Pack, so M-sheet column numbers must not be reused.
    """
    text_headers = [_cell_text(header) for header in headers]
    header_index = _build_header_index(text_headers)
    measures_col = header_index.get("measures")
    if measures_col is None:
        measures_col = _find_measures_column(text_headers)

    numbered = _mth_cols_from_numbered_headers(text_headers)
    if numbered:
        mth_cols = _remap_mth_cols_rightmost_first(numbered)
    else:
        date_cols = _mth_date_columns_from_headers(headers)
        ordered = sorted(date_cols, key=lambda item: item[0])
        mth_cols = {
            f"MTH{rank}": col_idx
            for rank, (col_idx, _label) in enumerate(reversed(ordered), start=1)
        }

    calendar = _calendar_from_header_cells(headers)
    if not calendar:
        date_cols = _mth_date_columns_from_headers(headers)
        if date_cols:
            calendar = _build_mth_calendar_from_dates(date_cols)

    return measures_col, mth_cols, calendar


def _metrics_from_vt_row(
    row: tuple,
    *,
    mth_cols: dict[str, int],
    calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
) -> dict[str, object]:
    """Compute RAW-equivalent metric values from one VTSTACK data row."""
    metrics: dict[str, object] = {}
    if measures_col:
        metrics["MEASURES"] = _cell_at_row(row, measures_col)

    for label, col_idx in mth_cols.items():
        metrics[label] = _cell_at_row(row, col_idx)

    mat1 = sum(_num(metrics.get(f"MTH{i}")) for i in range(1, 13))
    mat13 = sum(_num(metrics.get(f"MTH{i}")) for i in range(13, 25))
    metrics["MAT1"] = mat1 if mat1 else None
    metrics["MAT13"] = mat13 if mat13 else None

    anchor = calendar.get("MTH1")
    if anchor:
        anchor_year, anchor_month = anchor
        ytd1 = 0.0
        ytd13 = 0.0
        for mth_label in sorted(calendar, key=lambda k: int(k[3:])):
            year, month = calendar[mth_label]
            val = _num(metrics.get(mth_label))
            if year == anchor_year and 1 <= month <= anchor_month:
                ytd1 += val
            if year == anchor_year - 1 and 1 <= month <= anchor_month:
                ytd13 += val
        metrics["YTD1"] = ytd1 if ytd1 else None
        metrics["YTD13"] = ytd13 if ytd13 else None
    else:
        metrics["YTD1"] = None
        metrics["YTD13"] = None

    return metrics


def _resolve_table1_context(
    wb,
) -> tuple[
    str,
    dict[str, int],
    dict[str, tuple[int, int]],
    int | None,
    int | None,
    int | None,
    int | None,
    int | None,
    dict[tuple[str, str], str],
    dict[tuple[str, str], str],
    int,
]:
    """Return vtstack sheet, mth_cols, calendar, measures/brick/pack/market/type cols, maps."""
    raw_title = _truncate_sheet_name("RAW FOR NEXT STAGE")
    if raw_title not in wb.sheetnames:
        raise RuntimeError(f"No {raw_title!r} sheet in workbook")
    if "VTSTACK" not in wb.sheetnames:
        raise RuntimeError("No VTSTACK sheet in workbook")

    raw_row_count = max(wb[raw_title].max_row - 1, 0)

    vt_header = [cell.value for cell in wb["VTSTACK"][1]]
    measures_col, mth_cols, calendar = _column_layout_from_vtstack_headers(vt_header)
    vt_index = _build_header_index([_cell_text(v) for v in vt_header])
    brick_col = vt_index.get("brick")
    pack_col = vt_index.get("pack")
    market_col = vt_index.get("market")
    type_col = vt_index.get("type")

    o_sheet = _find_sheet_by_prefix(wb.sheetnames, "O")
    c_sheet = _find_sheet_by_prefix(wb.sheetnames, "C")
    pack1_map = _build_pack_map_for_sheet(wb[o_sheet], sheet_name=o_sheet) if o_sheet else {}
    pack2_map = _build_pack_map_for_sheet(wb[c_sheet], sheet_name=c_sheet) if c_sheet else {}

    if not calendar and mth_cols:
        logger.warning("TABLE 1 — YTD values may be empty (no MTH date calendar on VTSTACK)")

    return (
        "VTSTACK",
        mth_cols,
        calendar,
        measures_col,
        brick_col,
        pack_col,
        market_col,
        type_col,
        pack1_map,
        pack2_map,
        raw_row_count,
    )


def _parse_period_year_month(label: str) -> tuple[int, int] | None:
    match = PERIOD_DATE_HEADER.match(_cell_text(label))
    if not match:
        return None
    parts = _cell_text(label).split()[-1].split("/")
    if len(parts) != 2:
        return None
    return int(parts[0]), int(parts[1])


def _build_mth_calendar_from_dates(
    date_cols: list[tuple[int, str]],
) -> dict[str, tuple[int, int]]:
    """Map MTH1 (rightmost) … to (year, month) from MTH date-style headers."""
    calendar: dict[str, tuple[int, int]] = {}
    mth_dates = [
        (col_idx, label)
        for col_idx, label in date_cols
        if MTH_DATE_HEADER.match(_cell_text(label))
    ]
    ordered = sorted(mth_dates, key=lambda item: item[0])
    for rank, (_col_idx, old_label) in enumerate(reversed(ordered), start=1):
        year_month = _parse_period_year_month(old_label)
        if year_month:
            calendar[f"MTH{rank}"] = year_month
    return calendar


def _truncate_sheet_name(name: str) -> str:
    return name[:EXCEL_SHEET_NAME_MAX]


def _find_sheet_by_prefix(sheet_names: list[str], prefix: str) -> str | None:
    needle = f"{prefix.upper()}-"
    for name in sheet_names:
        if name.upper().startswith(needle):
            return name
    return None


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _write_mth_header_labels(
    ws, row_idx: int, mth_cols: dict[str, int]
) -> None:
    for label, col_idx in mth_cols.items():
        ws.cell(row_idx, col_idx, label)


def _remap_mth_cols_rightmost_first(mth_cols: dict[str, int]) -> dict[str, int]:
    """Assign MTH1 to the rightmost physical column, then MTH2, … leftward."""
    if not mth_cols:
        return {}
    by_position = sorted(mth_cols.items(), key=lambda item: item[1])
    return {
        f"MTH{rank}": col_idx
        for rank, (_old_label, col_idx) in enumerate(reversed(by_position), start=1)
    }


def _mth_date_columns_from_headers(
    headers: list,
) -> list[tuple[int, str]]:
    return [
        (col_idx, text)
        for col_idx, header in enumerate(headers, start=1)
        if (text := _cell_text(header)) and MTH_DATE_HEADER.match(text)
    ]


def _find_best_mth_date_columns(
    ws,
) -> tuple[int | None, list[tuple[int, str]]]:
    """Return the row with the most MTH yyyy/mm headers and those columns."""
    max_scan = min(20, ws.max_row + 1)
    best_row: int | None = None
    best_cols: list[tuple[int, str]] = []

    for row_idx in range(1, max_scan):
        headers = [
            ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)
        ]
        mth_dates = _mth_date_columns_from_headers(headers)
        if len(mth_dates) > len(best_cols):
            best_row = row_idx
            best_cols = mth_dates

    return best_row, best_cols


def _collect_period_columns(
    headers: list,
) -> tuple[dict[str, int], list[tuple[int, str]]]:
    """Return numbered MTH cols and date-style cols from a header row."""
    mth_cols: dict[str, int] = {}
    date_cols: list[tuple[int, str]] = []
    for col_idx, header in enumerate(headers, start=1):
        text = _cell_text(header)
        if not text:
            continue
        if MTH_HEADER.match(text):
            mth_cols[text.upper()] = col_idx
        elif PERIOD_DATE_HEADER.match(text):
            date_cols.append((col_idx, text))
    return mth_cols, date_cols


def _apply_period_column_labels(
    ws, row_idx: int, date_cols: list[tuple[int, str]]
) -> dict[str, int]:
    """Write MTH1, MTH2, … with the rightmost MTH date column as MTH1."""
    mth_cols: dict[str, int] = {}
    ordered = sorted(date_cols, key=lambda item: item[0])
    for rank, (col_idx, old_label) in enumerate(reversed(ordered), start=1):
        label = f"MTH{rank}"
        mth_cols[label] = col_idx
        ws.cell(row_idx, col_idx, label)
        logger.debug(
            "Renamed column %r → %s (col %d)", old_label, label, col_idx
        )
    if mth_cols:
        labels = ", ".join(
            f"{get_column_letter(col_idx)}={label}"
            for label, col_idx in sorted(mth_cols.items(), key=lambda x: x[1])
        )
        logger.info(
            "Renamed %d period column(s) on row %d (rightmost → MTH1): %s",
            len(mth_cols),
            row_idx,
            labels,
        )
    return mth_cols


def _scan_sheet_for_period_headers(
    ws,
) -> tuple[int | None, int | None, dict[str, int], list[tuple[int, str]]]:
    """Locate the header row with the most MTH/MAT date columns."""
    max_scan = min(20, ws.max_row + 1)
    best: tuple[int, int, int | None, dict[str, int], list[tuple[int, str]]] | None = (
        None
    )

    for row_idx in range(1, max_scan):
        headers = [
            ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)
        ]
        mth_cols, date_cols = _collect_period_columns(headers)
        if not date_cols and not mth_cols:
            continue

        measures_col: int | None = None
        for col_idx, header in enumerate(headers, start=1):
            if _cell_text(header).lower().startswith("measures"):
                measures_col = col_idx

        mth_date_cols = _mth_date_columns_from_headers(headers)

        # Prefer the row with the most MTH month date headers.
        score = len(mth_date_cols) * 100 + len(mth_cols)

        if best is None or score > best[0]:
            best = (score, row_idx, measures_col, mth_cols, date_cols)

    if best is None:
        return None, None, {}, []

    _, row_idx, measures_col, mth_cols, date_cols = best
    return row_idx, measures_col, mth_cols, date_cols


def _resolve_mth_columns_on_sheet(
    ws,
    *,
    sheet_name: str,
    header_row: int | None,
    mth_cols: dict[str, int],
    date_cols: list[tuple[int, str]],
) -> tuple[dict[str, int], dict[str, tuple[int, int]]]:
    """Build contiguous MTH1..MTHn map (MTH-only) and a YTD calendar."""
    mth_date_cols = [
        item for item in date_cols if MTH_DATE_HEADER.match(_cell_text(item[1]))
    ]
    if not mth_date_cols:
        _, mth_date_cols = _find_best_mth_date_columns(ws)

    if mth_date_cols and header_row is not None:
        logger.info(
            "Sheet %r — renaming %d MTH date column(s) to MTH1..MTH%d "
            "(rightmost = MTH1, MAT columns unchanged)",
            sheet_name,
            len(mth_date_cols),
            len(mth_date_cols),
        )
        calendar = _build_mth_calendar_from_dates(mth_date_cols)
        resolved = _apply_period_column_labels(ws, header_row, mth_date_cols)
        return resolved, calendar

    if mth_cols and header_row is not None:
        resolved = _remap_mth_cols_rightmost_first(mth_cols)
        _write_mth_header_labels(ws, header_row, resolved)
        logger.info(
            "Sheet %r — remapped %d numbered month column(s) to MTH1..MTH%d "
            "(rightmost = MTH1): %s",
            sheet_name,
            len(mth_cols),
            len(resolved),
            ", ".join(
                f"{get_column_letter(col)}={label}"
                for label, col in sorted(resolved.items(), key=lambda x: x[1])
            ),
        )
        _, mth_date_cols = _find_best_mth_date_columns(ws)
        calendar = (
            _build_mth_calendar_from_dates(mth_date_cols) if mth_date_cols else {}
        )
        return resolved, calendar

    return {}, {}


def _rename_period_columns_on_sheet(
    ws, *, sheet_name: str
) -> tuple[int | None, dict[str, int], dict[str, tuple[int, int]]]:
    """Rename MTH month headers; return (measures_col, {MTHn: col}, calendar)."""
    row_idx, measures_col, mth_cols, date_cols = _scan_sheet_for_period_headers(ws)
    if row_idx is None:
        return None, {}, {}

    resolved, calendar = _resolve_mth_columns_on_sheet(
        ws,
        sheet_name=sheet_name,
        header_row=row_idx,
        mth_cols=mth_cols,
        date_cols=date_cols,
    )
    return measures_col, resolved, calendar


def _convert_xls_to_xlsx(xls_path: Path) -> Path:
    xlsx_path = xls_path.with_suffix(".xlsx")
    workbook = pd.ExcelFile(xls_path, engine="xlrd")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name in workbook.sheet_names:
            frame = pd.read_excel(workbook, sheet_name=sheet_name, header=None)
            frame.to_excel(
                writer,
                sheet_name=_truncate_sheet_name(sheet_name),
                index=False,
                header=False,
            )
    xls_path.unlink(missing_ok=True)
    return xlsx_path


def _copy_worksheet(source_ws, target_wb: Workbook, title: str) -> None:
    title = _truncate_sheet_name(title)
    if title in target_wb.sheetnames:
        del target_wb[title]
    target_ws = target_wb.create_sheet(title)
    for row in source_ws.iter_rows():
        for cell in row:
            target_ws[cell.coordinate].value = cell.value


def _merge_workbook_files(paths: list[Path], destination: Path) -> Path:
    merged = Workbook()
    if merged.sheetnames:
        merged.remove(merged.active)

    for path in paths:
        normalized = path
        if path.suffix.lower() == ".xls":
            normalized = _convert_xls_to_xlsx(path)
        source_wb = load_workbook(normalized, data_only=False)
        for sheet_name in source_wb.sheetnames:
            _copy_worksheet(source_wb[sheet_name], merged, sheet_name)
        source_wb.close()
        if normalized != path and normalized.exists():
            normalized.unlink(missing_ok=True)

    merged.save(destination)
    merged.close()
    return destination


def _normalize_download_to_xlsx(path: Path) -> Path:
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        return path

    if suffix == ".xls":
        return _convert_xls_to_xlsx(path)

    if suffix == ".zip":
        with tempfile.TemporaryDirectory() as tmp:
            work_dir = Path(tmp)
            with zipfile.ZipFile(path) as zf:
                zf.extractall(work_dir)
            excel_files = sorted(
                list(work_dir.rglob("*.xlsx")) + list(work_dir.rglob("*.xls"))
            )
            if not excel_files:
                raise RuntimeError(f"ZIP has no Excel files: {path}")
            out = path.with_suffix(".xlsx")
            if len(excel_files) == 1:
                single = excel_files[0]
                if single.suffix.lower() == ".xls":
                    converted = _convert_xls_to_xlsx(single)
                    shutil.move(converted, out)
                else:
                    shutil.move(single, out)
            else:
                _merge_workbook_files(excel_files, out)
            path.unlink(missing_ok=True)
            return out

    raise RuntimeError(f"Unsupported Excel download type: {path.name}")


def _find_data_start_row(ws) -> int:
    """First data row below the pivot header (period / Product / Brick row)."""
    header_row, _, _, _ = _scan_sheet_for_period_headers(ws)
    if header_row is not None:
        return header_row + 1

    for row_idx in range(1, min(20, ws.max_row + 1)):
        first = _cell_text(ws.cell(row_idx, 1).value).lower()
        if first in {"product", "brick", "market"}:
            return row_idx + 1

    return 5


def _find_header_row(ws) -> int | None:
    """Pivot column header row (Brick/Product + Measures + MTH…)."""
    header_row, _, _, _ = _scan_sheet_for_period_headers(ws)
    if header_row is not None:
        return header_row

    data_start = _find_data_start_row(ws)
    return data_start - 1 if data_start > 1 else None


def _scan_layout_from_row_tuples(
    row_tuples: list[tuple],
) -> tuple[int | None, list[str], dict[str, int], int | None]:
    """Locate header row in the first scanned rows of a read-only sheet."""
    best: tuple[int, int, int | None, dict[str, int], list[str]] | None = None

    for row_idx, row in enumerate(row_tuples, start=1):
        headers = [_cell_text(value) for value in row]
        if not any(headers):
            continue

        mth_cols: dict[str, int] = {}
        date_cols = 0
        mth_date_cols = 0
        measures_col: int | None = None
        for col_idx, header in enumerate(headers, start=1):
            text = _cell_text(header)
            if not text:
                continue
            if MTH_HEADER.match(text):
                mth_cols[text.upper()] = col_idx
            elif PERIOD_DATE_HEADER.match(text):
                date_cols += 1
                if MTH_DATE_HEADER.match(text):
                    mth_date_cols += 1
            if text.lower().startswith("measures"):
                measures_col = col_idx

        if not date_cols and not mth_cols:
            continue

        score = mth_date_cols * 100 + len(mth_cols)
        if best is None or score > best[0]:
            best = (score, row_idx, measures_col, mth_cols, headers)

    if best is None:
        return None, [], {}, None

    _, row_idx, measures_col, _mth_cols, headers = best
    col_end = len(headers)
    while col_end > 0 and not headers[col_end - 1]:
        col_end -= 1
    headers = headers[:col_end]
    if measures_col is None:
        measures_col = _find_measures_column(headers)
    return row_idx, headers, _build_header_index(headers), measures_col


def _read_sheet_prefix_rows(ws, *, max_rows: int = 20) -> list[tuple]:
    rows: list[tuple] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx > max_rows:
            break
        rows.append(tuple(row))
    return rows


def _build_vtstack_col_spec(
    dest_headers: list[str],
    src_header_index: dict[str, int],
    *,
    market_value: str,
) -> list[str | int | None]:
    """Per-column mapping: 'market', 'type', source col (1-based), or None."""
    spec: list[str | int | None] = []
    for header in dest_headers:
        key = _normalize_header_key(header)
        if key == "type":
            spec.append("type")
        elif key == "market" and market_value:
            spec.append("market")
        else:
            spec.append(src_header_index.get(key))
    return spec


def _map_row_tuple_to_vtstack(
    row: tuple,
    spec: list[str | int | None],
    *,
    market_value: str,
    row_type: str,
    pack_col: int | None = None,
    market_src_col: int | None = None,
) -> list:
    values: list = []
    for item in spec:
        if item == "type":
            values.append(row_type)
        elif item == "market":
            values.append(market_value)
        elif isinstance(item, int):
            col_idx = item - 1
            values.append(row[col_idx] if col_idx < len(row) else None)
        else:
            values.append(None)
    # M sheet has no Pack column — use its Market label as the M-pack in VTSTACK.
    if pack_col and not _cell_text(
        values[pack_col - 1] if pack_col - 1 < len(values) else None
    ).strip():
        if market_src_col:
            market_pack = _cell_text(
                row[market_src_col - 1] if market_src_col - 1 < len(row) else None
            ).strip()
            if market_pack:
                while len(values) < pack_col:
                    values.append(None)
                values[pack_col - 1] = market_pack
    return values


def _is_repeated_header_row_tuple(
    row: tuple,
    *,
    measures_col: int | None,
    header_keys: set[str],
) -> bool:
    if measures_col is None:
        return False

    first_cols = [
        _normalize_header_key(value)
        for value in row[: min(8, len(row))]
        if _cell_text(value)
    ]
    if not first_cols:
        return False

    if first_cols[0] in header_keys and sum(
        1 for cell in first_cols if cell in header_keys
    ) >= 2:
        return True

    measures_idx = measures_col - 1
    if measures_idx < len(row):
        measures_val = _normalize_header_key(row[measures_idx])
        if measures_val.startswith("measures"):
            return True
    return False


def _is_measure_data_row_tuple(row: tuple, measures_col: int | None) -> bool:
    if measures_col is None:
        return True
    measures_idx = measures_col - 1
    if measures_idx >= len(row):
        return False
    value = _normalize_header_key(row[measures_idx])
    return value in {"units", "values", "unit", "value"}


def _collect_vtstack_rows_readonly(
    xlsx_path: Path,
    *,
    m_sheet: str,
    o_sheet: str | None,
    c_sheet: str,
    reader=None,
) -> tuple[
    list[list],
    dict[str, int],
    dict[tuple[str, str], str],
    dict[tuple[str, str], str],
]:
    """Build VTSTACK rows and O/C pack maps in one read-only workbook pass."""
    close_reader = reader is None
    if reader is None:
        reader = load_workbook(xlsx_path, read_only=True, data_only=True)
    pack1_map: dict[tuple[str, str], str] = {}
    pack2_map: dict[tuple[str, str], str] = {}
    try:
        m_ws = reader[m_sheet]
        prefix_rows = _read_sheet_prefix_rows(m_ws)
        m_header_row, m_headers, _m_index, m_measures_col = _scan_layout_from_row_tuples(
            prefix_rows
        )
        if m_header_row is None or not m_headers:
            raise RuntimeError(f"Could not read header row on M sheet {m_sheet!r}")

        market_value = ""
        market_col = _build_header_index(m_headers).get("market")
        if market_col:
            for row in m_ws.iter_rows(min_row=m_header_row + 1, values_only=True):
                value = _cell_text(
                    row[market_col - 1] if market_col - 1 < len(row) else None
                )
                if value and _normalize_header_key(value) != "market":
                    market_value = value
                    break

        vtstack_headers = list(m_headers)
        if "pack" not in {_normalize_header_key(h) for h in vtstack_headers}:
            brick_idx = next(
                (
                    i
                    for i, h in enumerate(vtstack_headers)
                    if _normalize_header_key(h) == "brick"
                ),
                0,
            )
            vtstack_headers.insert(brick_idx, "Pack")
        if "TYPE" not in {_normalize_header_key(h) for h in vtstack_headers}:
            # Place TYPE after Pack so C/O competitor & own packs land in Pack2.
            pack_idx = next(
                (
                    i
                    for i, h in enumerate(vtstack_headers)
                    if _normalize_header_key(h) == "pack"
                ),
                None,
            )
            insert_at = (pack_idx + 1) if pack_idx is not None else 0
            vtstack_headers.insert(insert_at, "TYPE")

        vtstack_header_index = _build_header_index(vtstack_headers)
        vtstack_pack_col = vtstack_header_index.get("pack")

        vtstack_rows: list[list] = [vtstack_headers]
        row_counts: dict[str, int] = {}

        # C → O → M so TABLE 1 Pack2 walks competitor, own, then market packs.
        for source_sheet in (c_sheet, o_sheet, m_sheet):
            if not source_sheet:
                continue
            src_ws = reader[source_sheet]
            src_prefix = _read_sheet_prefix_rows(src_ws)
            src_header_row, src_headers, src_index, src_measures_col = (
                _scan_layout_from_row_tuples(src_prefix)
            )
            if src_header_row is None:
                logger.warning("Skipping %r — no header row found", source_sheet)
                continue

            fill_market = market_value if source_sheet != m_sheet else ""
            row_type = _vtstack_type_for_sheet(
                source_sheet, o_sheet=o_sheet, c_sheet=c_sheet, m_sheet=m_sheet
            )
            spec = _build_vtstack_col_spec(
                vtstack_headers, src_index, market_value=fill_market
            )
            header_keys = {_normalize_header_key(h) for h in src_headers if h}
            measures_col = src_measures_col or m_measures_col
            pack_col = src_index.get("pack")
            brick_col = src_index.get("brick")
            copied = 0

            for row in src_ws.iter_rows(min_row=src_header_row + 1, values_only=True):
                row_tuple = tuple(row)
                if _is_repeated_header_row_tuple(
                    row_tuple,
                    measures_col=measures_col,
                    header_keys=header_keys,
                ):
                    continue
                if not _is_measure_data_row_tuple(row_tuple, measures_col):
                    continue
                if source_sheet == o_sheet:
                    _add_pack_map_entry(
                        pack1_map,
                        row_tuple,
                        pack_col=pack_col,
                        brick_col=brick_col,
                        measures_col=measures_col,
                    )
                elif source_sheet == c_sheet:
                    _add_pack_map_entry(
                        pack2_map,
                        row_tuple,
                        pack_col=pack_col,
                        brick_col=brick_col,
                        measures_col=measures_col,
                    )
                vtstack_rows.append(
                    _map_row_tuple_to_vtstack(
                        row_tuple,
                        spec,
                        market_value=fill_market,
                        row_type=row_type,
                        pack_col=vtstack_pack_col,
                        market_src_col=src_index.get("market"),
                    )
                )
                copied += 1
            row_counts[source_sheet] = copied
            logger.info(
                "VTSTACK — stacked %d row(s) from %r (TYPE=%s)",
                copied,
                source_sheet,
                row_type,
            )

        if pack1_map:
            logger.info("Pack map — %r: %d brick+measure rows", o_sheet, len(pack1_map))
        if pack2_map:
            logger.info("Pack map — %r: %d brick+measure rows", c_sheet, len(pack2_map))

        if len(vtstack_rows) <= 1:
            raise RuntimeError("VTSTACK has no source rows to stack")

        logger.info(
            "VTSTACK rows collected — %s (%d columns, market=%r)",
            ", ".join(f"{name}={count}" for name, count in row_counts.items()),
            len(vtstack_headers),
            market_value or "(from M rows)",
        )
        return vtstack_rows, row_counts, pack1_map, pack2_map
    finally:
        if close_reader:
            reader.close()


def _normalize_header_key(text: str) -> str:
    return _cell_text(text).lower()


def _read_header_cells(ws, header_row: int, col_end: int) -> list[str]:
    return [
        _cell_text(ws.cell(header_row, col).value) for col in range(1, col_end + 1)
    ]


def _build_header_index(headers: list[str]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col_idx, header in enumerate(headers, start=1):
        key = _normalize_header_key(header)
        if key:
            index[key] = col_idx
    return index


def _find_measures_column(headers: list[str]) -> int | None:
    for col_idx, header in enumerate(headers, start=1):
        if _normalize_header_key(header).startswith("measures"):
            return col_idx
    return None


def _scan_sheet_layout(ws) -> tuple[int | None, list[str], dict[str, int], int | None]:
    """Return (header_row, headers, header_index, measures_col) for a pivot sheet."""
    header_row, measures_col, _, _ = _scan_sheet_for_period_headers(ws)
    if header_row is None:
        header_row = _find_header_row(ws)
    if header_row is None:
        return None, [], {}, None

    col_end = max(ws.max_column, 1)
    headers = _read_header_cells(ws, header_row, col_end)
    if measures_col is None:
        measures_col = _find_measures_column(headers)
    return header_row, headers, _build_header_index(headers), measures_col


def _apply_auto_filter(ws) -> None:
    """Enable Excel AutoFilter dropdowns on the detected header row through last data row."""
    if ws.max_row < 1 or ws.max_column < 1:
        return
    header_row, headers, _, _ = _scan_sheet_layout(ws)
    if header_row is None:
        header_row = _find_header_row(ws)
    if header_row is None:
        header_row = 1
    end_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A{header_row}:{end_col}{ws.max_row}"


def _apply_auto_filters_to_workbook(wb) -> None:
    for sheet_name in wb.sheetnames:
        _apply_auto_filter(wb[sheet_name])
    logger.info("Applied AutoFilter dropdowns on %d sheet(s)", len(wb.sheetnames))


def _add_vtstack_sheet(
    wb,
    *,
    vtstack_rows: list[list],
) -> str:
    """Write pre-built VTSTACK rows into the workbook."""
    sheet_title = _truncate_sheet_name("VTSTACK")
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(sheet_title)
    for row in vtstack_rows:
        ws.append(row)
    logger.info(
        "Added VTSTACK sheet — %d total row(s) including header",
        len(vtstack_rows),
    )
    return sheet_title


def _quote_sheet_name(sheet_name: str) -> str:
    if re.search(r"[^A-Za-z0-9_]", sheet_name):
        return f"'{sheet_name}'"
    return sheet_name


def _cell_ref(sheet_name: str, col_idx: int, row: int) -> str:
    return f"{_quote_sheet_name(sheet_name)}!{get_column_letter(col_idx)}{row}"


def _mth_range_sum_formula(
    sheet_name: str,
    mth_cols: dict[str, int],
    *,
    row: int,
    mth_start: int,
    mth_end: int,
) -> str:
    refs: list[str] = []
    for mth_n in range(mth_start, mth_end + 1):
        col_idx = mth_cols.get(f"MTH{mth_n}")
        if col_idx:
            refs.append(f"{_quote_sheet_name(sheet_name)}!{get_column_letter(col_idx)}{row}")
    if not refs:
        return ""
    if len(refs) == 1:
        return f"={refs[0]}"
    return f"=SUM({','.join(refs)})"


def _ytd_sum_formula(
    sheet_name: str,
    mth_cols: dict[str, int],
    calendar: dict[str, tuple[int, int]],
    *,
    row: int,
    year_offset: int,
) -> str:
    """YTD: sum Jan → anchor month (from MTH1/rightmost column date, any month)."""
    anchor = calendar.get("MTH1")
    if not anchor:
        return ""
    anchor_year, anchor_month = anchor
    target_year = anchor_year + year_offset

    refs: list[str] = []
    sheet = _quote_sheet_name(sheet_name)
    for mth_label in sorted(calendar, key=lambda k: int(k[3:])):
        year, month = calendar[mth_label]
        # Include only months Jan..anchor_month for the target calendar year.
        if year != target_year or month < 1 or month > anchor_month:
            continue
        col_idx = mth_cols.get(mth_label)
        if col_idx:
            refs.append(f"{sheet}!{get_column_letter(col_idx)}{row}")

    if not refs:
        return ""
    if len(refs) == 1:
        return f"={refs[0]}"
    return f"=SUM({','.join(refs)})"


def _add_raw_for_next_stage_sheet(
    wb,
    *,
    vtstack_data_rows: list[list],
    mth_cols: dict[str, int],
    mth_calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
) -> str:
    """RAW FOR NEXT STAGE — computed values (same numbers as formulas, much smaller file)."""
    sheet_title = _truncate_sheet_name("RAW FOR NEXT STAGE")
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]

    if mth_calendar.get("MTH1"):
        year, month = mth_calendar["MTH1"]
        logger.info(
            "YTD anchor — rightmost MTH1 = %04d/%02d → "
            "YTD1: Jan–%02d %d, YTD13: Jan–%02d %d",
            year,
            month,
            month,
            year,
            month,
            year - 1,
        )

    ws = wb.create_sheet(sheet_title)
    ws.append(RAW_FOR_NEXT_STAGE_HEADERS)
    for vt_row in vtstack_data_rows:
        metrics = _metrics_from_vt_row(
            tuple(vt_row),
            mth_cols=mth_cols,
            calendar=mth_calendar,
            measures_col=measures_col,
        )
        ws.append([metrics.get(header) for header in RAW_FOR_NEXT_STAGE_HEADERS])

    rows_written = len(vtstack_data_rows)
    logger.info(
        "Added RAW FOR NEXT STAGE (%d headers, %d data rows, values only)",
        len(RAW_FOR_NEXT_STAGE_HEADERS),
        rows_written,
    )
    return sheet_title


TABLE1_CSV_BATCH_SIZE = CSV_BATCH_SIZE


def _write_csv_rows(csv_path: Path, rows: Iterator[list], *, label: str) -> int:
    """Stream rows to CSV with batching; return data row count (excludes header)."""
    csv_path = Path(csv_path)
    rows_written = 0
    batch: list[list] = []
    with csv_path.open(
        "w", newline="", encoding="utf-8", buffering=8 * 1024 * 1024
    ) as handle:
        writer = csv.writer(handle)
        for row in rows:
            batch.append(["" if value is None else value for value in row])
            rows_written += 1
            if len(batch) >= CSV_BATCH_SIZE:
                writer.writerows(batch)
                batch.clear()
        if batch:
            writer.writerows(batch)

    data_rows = max(rows_written - 1, 0)
    size_mb = csv_path.stat().st_size / 1024 / 1024
    logger.info("Exported %s to %s (%d data rows, %.1f MB)", label, csv_path.name, data_rows, size_mb)
    return data_rows


def _export_vtstack_csv(csv_path: Path, vtstack_rows: list[list]) -> Path:
    def _rows() -> Iterator[list]:
        for row in vtstack_rows:
            yield list(row)

    _write_csv_rows(csv_path, _rows(), label="VTSTACK")
    return csv_path


def _export_raw_csv(
    csv_path: Path,
    vtstack_data_rows: list[list],
    *,
    mth_cols: dict[str, int],
    mth_calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
) -> Path:
    def _rows() -> Iterator[list]:
        yield list(RAW_FOR_NEXT_STAGE_HEADERS)
        for vt_row in vtstack_data_rows:
            metrics = _metrics_from_vt_row(
                tuple(vt_row),
                mth_cols=mth_cols,
                calendar=mth_calendar,
                measures_col=measures_col,
            )
            yield [metrics.get(header) for header in RAW_FOR_NEXT_STAGE_HEADERS]

    _write_csv_rows(csv_path, _rows(), label="RAW FOR NEXT STAGE")
    return csv_path


def _table1_base_row(
    row_tuple: tuple,
    *,
    mth_cols: dict[str, int],
    mth_calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
    brick_col: int | None,
    pack_col: int | None,
    market_col: int | None = None,
    type_col: int | None = None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
) -> list:
    """Pack1 (O), Pack2 (C/O/M pack), Brick, Type, MEASURES, MTHs."""
    metrics = _metrics_from_vt_row(
        row_tuple,
        mth_cols=mth_cols,
        calendar=mth_calendar,
        measures_col=measures_col,
    )
    measures_text = _cell_text(metrics.get("MEASURES")).strip()
    brick_raw = _cell_at_row(row_tuple, brick_col) if brick_col else None
    brick_code = _brick_code_only(brick_raw) if brick_raw is not None else None
    brick_display = _cell_text(brick_raw).strip() or None

    key = (brick_code or "", measures_text.lower())
    pack1 = pack1_map.get(key) if brick_code and measures_text else None
    pack2 = _table1_pack2(
        row_tuple=row_tuple,
        key=key,
        type_col=type_col,
        pack_col=pack_col,
        market_col=market_col,
        pack1_map=pack1_map,
        pack2_map=pack2_map,
    )
    row_type = _table1_type_label(pack1=pack1, pack2=pack2)

    return [pack1, pack2, brick_display, row_type, metrics.get("MEASURES")] + [
        metrics.get(header) for header in ATTRIBUTE_VALUE_MTH_HEADERS
    ]


def _iter_table1_rows_from_data(
    vtstack_data_rows: list[list],
    *,
    mth_cols: dict[str, int],
    mth_calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
    brick_col: int | None,
    pack_col: int | None,
    market_col: int | None = None,
    type_col: int | None = None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
    include_header: bool = True,
) -> Iterator[list]:
    """Yield TABLE 1 rows — 6 Attribute/Value lines per VTSTACK row."""
    if include_header:
        yield list(ATTRIBUTE_VALUE_HEADERS)

    for vt_row in vtstack_data_rows:
        row_tuple = tuple(vt_row)
        metrics = _metrics_from_vt_row(
            row_tuple,
            mth_cols=mth_cols,
            calendar=mth_calendar,
            measures_col=measures_col,
        )
        base = _table1_base_row(
            row_tuple,
            mth_cols=mth_cols,
            mth_calendar=mth_calendar,
            measures_col=measures_col,
            brick_col=brick_col,
            pack_col=pack_col,
            market_col=market_col,
            type_col=type_col,
            pack1_map=pack1_map,
            pack2_map=pack2_map,
        )
        for attribute in ATTRIBUTE_VALUE_ATTRIBUTE_LABELS:
            yield base + [attribute, metrics.get(attribute)]


def _write_table1_csv(
    csv_path: Path,
    *,
    vtstack_data_rows: list[list] | None = None,
    vt_ws=None,
    mth_cols: dict[str, int],
    mth_calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
    brick_col: int | None,
    pack_col: int | None,
    market_col: int | None = None,
    type_col: int | None = None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
    raw_row_count: int | None = None,
) -> int:
    """Write TABLE 1 rows to a sidecar CSV; return data row count."""
    csv_path = Path(csv_path)
    if vtstack_data_rows is not None:
        row_iter = _iter_table1_rows_from_data(
            vtstack_data_rows,
            mth_cols=mth_cols,
            mth_calendar=mth_calendar,
            measures_col=measures_col,
            brick_col=brick_col,
            pack_col=pack_col,
            market_col=market_col,
            type_col=type_col,
            pack1_map=pack1_map,
            pack2_map=pack2_map,
        )
    elif vt_ws is not None:
        limit = raw_row_count or max(vt_ws.max_row - 1, 0)

        def _sheet_rows() -> Iterator[list]:
            yield list(ATTRIBUTE_VALUE_HEADERS)
            for row_idx, row in enumerate(
                vt_ws.iter_rows(min_row=2, values_only=True), start=2
            ):
                if row_idx > limit + 1:
                    break
                yield from _iter_table1_rows_from_data(
                    [list(row)],
                    mth_cols=mth_cols,
                    mth_calendar=mth_calendar,
                    measures_col=measures_col,
                    brick_col=brick_col,
                    pack_col=pack_col,
                    market_col=market_col,
                    type_col=type_col,
                    pack1_map=pack1_map,
                    pack2_map=pack2_map,
                    include_header=False,
                )

        row_iter = _sheet_rows()
    else:
        raise ValueError("Provide vtstack_data_rows or vt_ws for TABLE 1 export")

    rows_written = 0
    batch: list[list] = []
    with csv_path.open(
        "w", newline="", encoding="utf-8", buffering=8 * 1024 * 1024
    ) as handle:
        writer = csv.writer(handle)
        for row in row_iter:
            batch.append(["" if value is None else value for value in row])
            rows_written += 1
            if len(batch) >= CSV_BATCH_SIZE:
                writer.writerows(batch)
                batch.clear()
        if batch:
            writer.writerows(batch)

    data_rows = max(rows_written - 1, 0)
    size_mb = csv_path.stat().st_size / 1024 / 1024
    logger.info(
        "Exported TABLE 1 to %s (%d data rows, %.1f MB)",
        csv_path.name,
        data_rows,
        size_mb,
    )
    return data_rows


def _add_table1_sheet(
    wb,
    vtstack_data_rows: list[list],
    *,
    mth_cols: dict[str, int],
    mth_calendar: dict[str, tuple[int, int]],
    measures_col: int | None,
    brick_col: int | None,
    pack_col: int | None,
    market_col: int | None = None,
    type_col: int | None = None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
) -> str:
    """Write TABLE 1 (Attribute/Value) into the workbook."""
    sheet_title = _truncate_sheet_name(ATTRIBUTE_VALUE_SHEET_TITLE)
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(sheet_title)
    rows_written = 0
    for row in _iter_table1_rows_from_data(
        vtstack_data_rows,
        mth_cols=mth_cols,
        mth_calendar=mth_calendar,
        measures_col=measures_col,
        brick_col=brick_col,
        pack_col=pack_col,
        market_col=market_col,
        type_col=type_col,
        pack1_map=pack1_map,
        pack2_map=pack2_map,
    ):
        ws.append(["" if value is None else value for value in row])
        rows_written += 1
    logger.info(
        "Added TABLE 1 sheet — %d rows (Attribute/Value format)",
        rows_written,
    )
    return sheet_title


def _is_iqvia_pivot_sheet(sheet_name: str) -> bool:
    """True for raw C/O/M designer sheets (not VTSTACK / TABLE 1 / RAW)."""
    name = sheet_name.strip().upper()
    return name.startswith("C-") or name.startswith("O-") or name.startswith("M-")


def _combined_csv_start_row(ws, sheet_name: str) -> int:
    """First row to write for combined CSV — skip IQVIA pivot preamble on C/O/M."""
    if not _is_iqvia_pivot_sheet(sheet_name):
        return 1
    prefix_rows = _read_sheet_prefix_rows(ws)
    header_row, _headers, _, _ = _scan_layout_from_row_tuples(prefix_rows)
    if header_row is not None:
        return header_row
    fallback = _find_header_row(ws)
    return fallback if fallback is not None else 1


def export_workbook_to_combined_csv(xlsx_path: Path) -> Path:
    """
    Export every workbook sheet into one CSV file.

    First column is ``Sheet`` (sheet name); remaining columns are that row's values.
    C/O/M pivot sheets start at the detected header row (Product/Pack/Brick/Measures…)
    so PivotTable / filter / MAT preamble rows are not included.
    """
    xlsx_path = Path(xlsx_path)
    csv_path = combined_csv_path_for(xlsx_path)
    xlsx_mb = xlsx_path.stat().st_size / 1024 / 1024
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows_written = 0
    skipped_preamble = 0
    sheet_count = len(wb.sheetnames)
    batch: list[list] = []
    try:
        with csv_path.open(
            "w", newline="", encoding="utf-8", buffering=8 * 1024 * 1024
        ) as handle:
            writer = csv.writer(handle)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                start_row = _combined_csv_start_row(ws, sheet_name)
                if start_row > 1:
                    skipped_preamble += start_row - 1
                    logger.info(
                        "CSV export %r — skipping %d preamble row(s), starting at header row %d",
                        sheet_name,
                        start_row - 1,
                        start_row,
                    )
                for row in ws.iter_rows(min_row=start_row, values_only=True):
                    batch.append(
                        [sheet_name]
                        + ["" if value is None else value for value in row]
                    )
                    rows_written += 1
                    if len(batch) >= CSV_BATCH_SIZE:
                        writer.writerows(batch)
                        batch.clear()
            if batch:
                writer.writerows(batch)
    finally:
        wb.close()
    size_mb = csv_path.stat().st_size / 1024 / 1024
    logger.info(
        "Exported combined CSV %s — %d rows (%d preamble rows skipped), "
        "%d sheet(s), %.1f MB (xlsx %.1f MB)",
        csv_path.name,
        rows_written,
        skipped_preamble,
        sheet_count,
        size_mb,
        xlsx_mb,
    )
    return csv_path


def export_postprocess_csvs(
    xlsx_path: Path,
    *,
    vtstack_rows: list[list],
    mth_cols: dict[str, int] | None = None,
    mth_calendar: dict[str, tuple[int, int]] | None = None,
    measures_col: int | None = None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
) -> Path:
    """Write VTSTACK, RAW, and TABLE 1 sidecar CSVs; return TABLE 1 path."""
    xlsx_path = Path(xlsx_path)
    vtstack_csv = vtstack_csv_path_for(xlsx_path)
    raw_csv = raw_csv_path_for(xlsx_path)
    table1_csv = table1_csv_path_for(xlsx_path)

    vt_measures_col, vt_mth_cols, vt_calendar = _column_layout_from_vtstack_headers(
        vtstack_rows[0]
    )
    measures_col = measures_col or vt_measures_col
    mth_cols = mth_cols or vt_mth_cols
    mth_calendar = mth_calendar or vt_calendar
    vt_index = _build_header_index([_cell_text(v) for v in vtstack_rows[0]])

    _export_vtstack_csv(vtstack_csv, vtstack_rows)
    _export_raw_csv(
        raw_csv,
        vtstack_rows[1:],
        mth_cols=mth_cols,
        mth_calendar=mth_calendar,
        measures_col=measures_col,
    )
    _write_table1_csv(
        table1_csv,
        vtstack_data_rows=vtstack_rows[1:],
        mth_cols=mth_cols,
        mth_calendar=mth_calendar,
        measures_col=measures_col,
        brick_col=vt_index.get("brick"),
        pack_col=vt_index.get("pack"),
        market_col=vt_index.get("market"),
        type_col=vt_index.get("type"),
        pack1_map=pack1_map,
        pack2_map=pack2_map,
    )
    return table1_csv


def export_pivot_sheets_to_csv(xlsx_path: Path) -> list[Path]:
    """Export every sheet in the workbook to a sidecar CSV next to the xlsx.

    Returns the list of CSV paths written (one per sheet).
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    csv_paths: list[Path] = []
    try:
        for sheet_name in wb.sheetnames:
            safe_name = re.sub(r'[\\/*?:"<>|]', "_", sheet_name)
            csv_path = xlsx_path.with_name(f"{xlsx_path.stem}_{safe_name}.csv")
            rows_written = 0
            with csv_path.open(
                "w", newline="", encoding="utf-8", buffering=8 * 1024 * 1024
            ) as handle:
                writer = csv.writer(handle)
                for row in wb[sheet_name].iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
                    rows_written += 1
            size_mb = csv_path.stat().st_size / 1024 / 1024
            logger.info(
                "Exported sheet %r → %s (%d rows, %.1f MB)",
                sheet_name,
                csv_path.name,
                rows_written,
                size_mb,
            )
            csv_paths.append(csv_path)
    finally:
        wb.close()
    return csv_paths


def _save_workbook_atomic(wb, destination: Path) -> None:
    """Save workbook via temp file so an interrupted write cannot corrupt the original."""
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = destination.with_suffix(destination.suffix + ".tmp")
    try:
        wb.save(tmp_path)
        tmp_path.replace(destination)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise


def compress_for_delivery(
    path: Path, *, max_mb: float = MAX_WORKBOOK_MB
) -> Path:
    """
    Losslessly zip a CSV/XLSX when it exceeds max_mb (DEFLATE-9).

    Data is unchanged — unzip to get the identical original file.
    Plain CSV with Attribute/Value + all sheets cannot stay under 100 MB
    without compression.
    """
    path = Path(path)
    size_mb = path.stat().st_size / 1024 / 1024
    if size_mb <= max_mb:
        logger.info("%s (%.1f MB) — under %.0f MB limit", path.name, size_mb, max_mb)
        return path

    zip_path = path.with_suffix(".zip")
    if zip_path.exists():
        zip_path.unlink()
    logger.info(
        "%.1f MB > %.0f MB — zipping for upload (lossless, same CSV/XLSX inside)…",
        size_mb,
        max_mb,
    )
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        zf.write(path, path.name)
    original_name = path.name
    path.unlink()
    zip_mb = zip_path.stat().st_size / 1024 / 1024
    logger.info(
        "Delivery file: %s (%.1f MB, unzip → %s)",
        zip_path.name,
        zip_mb,
        original_name,
    )
    return zip_path


# Backward-compatible aliases
compress_workbook_for_delivery = compress_for_delivery
compress_workbook_if_oversized = compress_for_delivery


def _remove_table1_sheet_if_present(wb) -> bool:
    """Drop TABLE 1 / legacy ATTRIBUTE VALUE sheet from workbook if present."""
    removed = False
    for title in (ATTRIBUTE_VALUE_SHEET_TITLE, "ATTRIBUTE VALUE"):
        sheet_title = _truncate_sheet_name(title)
        if sheet_title in wb.sheetnames:
            del wb[sheet_title]
            removed = True
    return removed


def _export_table1_sidecar(
    wb,
    xlsx_path: Path,
    *,
    raw_row_count: int | None = None,
) -> Path:
    """Write TABLE 1 to a sidecar CSV next to the workbook."""
    (
        vtstack_sheet,
        mth_cols,
        mth_calendar,
        measures_col,
        brick_col,
        pack_col,
        market_col,
        type_col,
        pack1_map,
        pack2_map,
        resolved_raw_row_count,
    ) = _resolve_table1_context(wb)
    csv_path = table1_csv_path_for(xlsx_path)
    _write_table1_csv(
        csv_path,
        vt_ws=wb[vtstack_sheet],
        mth_cols=mth_cols,
        mth_calendar=mth_calendar,
        measures_col=measures_col,
        brick_col=brick_col,
        pack_col=pack_col,
        market_col=market_col,
        type_col=type_col,
        pack1_map=pack1_map,
        pack2_map=pack2_map,
        raw_row_count=(
            raw_row_count
            if raw_row_count is not None
            else resolved_raw_row_count
        ),
    )
    return csv_path


def _export_table1_sidecar_from_memory(
    xlsx_path: Path,
    vtstack_rows: list[list],
    *,
    mth_cols: dict[str, int] | None = None,
    mth_calendar: dict[str, tuple[int, int]] | None = None,
    measures_col: int | None = None,
    pack1_map: dict[tuple[str, str], str],
    pack2_map: dict[tuple[str, str], str],
) -> Path:
    """Write TABLE 1 CSV from VTSTACK rows already in memory."""
    vt_index = _build_header_index([_cell_text(v) for v in vtstack_rows[0]])
    brick_col = vt_index.get("brick")
    pack_col = vt_index.get("pack")
    market_col = vt_index.get("market")
    type_col = vt_index.get("type")
    vt_measures_col, vt_mth_cols, vt_calendar = _column_layout_from_vtstack_headers(
        vtstack_rows[0]
    )
    measures_col = measures_col or vt_measures_col
    mth_cols = mth_cols or vt_mth_cols
    calendar = mth_calendar or vt_calendar
    csv_path = table1_csv_path_for(xlsx_path)
    _write_table1_csv(
        csv_path,
        vtstack_data_rows=vtstack_rows[1:],
        mth_cols=mth_cols,
        mth_calendar=calendar,
        measures_col=measures_col,
        brick_col=brick_col,
        pack_col=pack_col,
        market_col=market_col,
        type_col=type_col,
        pack1_map=pack1_map,
        pack2_map=pack2_map,
    )
    return csv_path


def export_table1_csv_from_xlsx(path: Path) -> Path:
    """Export TABLE 1 to sidecar CSV and remove the sheet from the workbook."""
    xlsx_path = Path(path)
    vtstack_csv = vtstack_csv_path_for(xlsx_path)
    if vtstack_csv.is_file():
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            vtstack_rows: list[list] = []
            with vtstack_csv.open("r", encoding="utf-8", newline="") as handle:
                reader = csv.reader(handle)
                vtstack_rows = [row for row in reader]
            if not vtstack_rows:
                raise RuntimeError(f"VTSTACK CSV is empty: {vtstack_csv.name}")
        finally:
            wb.close()
        pack1_map: dict[tuple[str, str], str] = {}
        pack2_map: dict[tuple[str, str], str] = {}
        reader = load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            o_sheet = _find_sheet_by_prefix(sheet_names, "O")
            c_sheet = _find_sheet_by_prefix(sheet_names, "C")
            if o_sheet:
                pack1_map = _build_pack_map_for_sheet(reader[o_sheet], sheet_name=o_sheet)
            if c_sheet:
                pack2_map = _build_pack_map_for_sheet(reader[c_sheet], sheet_name=c_sheet)
        finally:
            reader.close()
        csv_path = _export_table1_sidecar_from_memory(
            xlsx_path,
            vtstack_rows,
            pack1_map=pack1_map,
            pack2_map=pack2_map,
        )
        wb = load_workbook(xlsx_path)
        if _remove_table1_sheet_if_present(wb):
            _save_workbook_atomic(wb, xlsx_path)
        wb.close()
        return csv_path

    raw_title = _truncate_sheet_name("RAW FOR NEXT STAGE")
    wb = load_workbook(xlsx_path)
    if raw_title not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            f"No {raw_title!r} sheet in {xlsx_path.name}; sheets: {wb.sheetnames}"
        )
    csv_path = _export_table1_sidecar(wb, xlsx_path)
    if _remove_table1_sheet_if_present(wb):
        _save_workbook_atomic(wb, xlsx_path)
        logger.info("Removed TABLE 1 sheet from %s (kept in sidecar CSV)", xlsx_path.name)
    wb.close()
    return csv_path


def postprocess_iqvia_xlsx(
    path: Path, *, product: str, zip_for_delivery: bool = True
) -> Path:
    """
    Normalize IQVIA Excel download:
    - Workbook: C/O/M + VTSTACK + RAW FOR NEXT STAGE + TABLE 1 (Attribute/Value)
    - One combined CSV with all sheets (first column = Sheet name)

    Returns the combined CSV path.
    """
    del product  # kept for caller compatibility
    xlsx_path = _normalize_download_to_xlsx(Path(path))

    reader = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = list(reader.sheetnames)
    c_sheet = _find_sheet_by_prefix(sheet_names, "C")
    o_sheet = _find_sheet_by_prefix(sheet_names, "O")
    m_sheet = _find_sheet_by_prefix(sheet_names, "M")
    if not c_sheet:
        reader.close()
        raise RuntimeError(
            f"No C sheet found in {xlsx_path.name}; sheets: {sheet_names}"
        )
    if not m_sheet:
        reader.close()
        raise RuntimeError(
            f"No M sheet found in {xlsx_path.name}; sheets: {sheet_names}"
        )

    logger.info("Building VTSTACK rows (read-only scan of %s)…", xlsx_path.name)
    vtstack_rows, _row_counts, pack1_map, pack2_map = _collect_vtstack_rows_readonly(
        xlsx_path,
        m_sheet=m_sheet,
        o_sheet=o_sheet,
        c_sheet=c_sheet,
        reader=reader,
    )
    reader.close()

    logger.info("Loading workbook to rename C/O/M period columns…")
    wb = load_workbook(xlsx_path)

    measures_col: int | None = None
    mth_cols: dict[str, int] = {}
    mth_calendar: dict[str, tuple[int, int]] = {}
    for sheet_name in (c_sheet, o_sheet, m_sheet):
        if not sheet_name:
            continue
        sheet_measures, sheet_mth, sheet_calendar = _rename_period_columns_on_sheet(
            wb[sheet_name], sheet_name=sheet_name
        )
        if sheet_name == m_sheet:
            measures_col = sheet_measures
            mth_cols = sheet_mth
            mth_calendar = sheet_calendar
        elif sheet_name == c_sheet and not mth_cols:
            measures_col = sheet_measures
            mth_cols = sheet_mth
        if not mth_calendar and sheet_calendar:
            mth_calendar = sheet_calendar

    if not mth_cols:
        wb.close()
        raise RuntimeError(
            f"Could not find MTH columns on sheet {c_sheet!r} in {xlsx_path.name}"
        )

    vtstack_sheet = _add_vtstack_sheet(wb, vtstack_rows=vtstack_rows)
    vt_index = _build_header_index([_cell_text(v) for v in vtstack_rows[0]])
    brick_col = vt_index.get("brick")
    pack_col = vt_index.get("pack")
    market_col = vt_index.get("market")
    type_col = vt_index.get("type")
    vt_measures_col, vt_mth_cols, vt_calendar = _column_layout_from_vtstack_headers(
        vtstack_rows[0]
    )

    _add_raw_for_next_stage_sheet(
        wb,
        vtstack_data_rows=vtstack_rows[1:],
        mth_cols=vt_mth_cols,
        mth_calendar=vt_calendar or mth_calendar,
        measures_col=vt_measures_col,
    )
    _add_table1_sheet(
        wb,
        vtstack_rows[1:],
        mth_cols=vt_mth_cols,
        mth_calendar=vt_calendar or mth_calendar,
        measures_col=vt_measures_col,
        brick_col=brick_col,
        pack_col=pack_col,
        market_col=market_col,
        type_col=type_col,
        pack1_map=pack1_map,
        pack2_map=pack2_map,
    )

    if not o_sheet:
        logger.warning("No O sheet found — VTSTACK will stack M then C only")

    _apply_auto_filters_to_workbook(wb)
    sheet_count = len(wb.sheetnames)
    _save_workbook_atomic(wb, xlsx_path)
    wb.close()

    combined_csv = export_workbook_to_combined_csv(xlsx_path)
    xlsx_mb = xlsx_path.stat().st_size / 1024 / 1024
    csv_mb = combined_csv.stat().st_size / 1024 / 1024
    logger.info(
        "Post-processed %s — %d sheets (%.1f MB) → %s (%.1f MB)",
        xlsx_path.name,
        sheet_count,
        xlsx_mb,
        combined_csv.name,
        csv_mb,
    )
    return compress_for_delivery(combined_csv) if zip_for_delivery else combined_csv
